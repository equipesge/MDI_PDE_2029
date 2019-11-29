from Sistema import Sistema;

from openpyxl import load_workbook;
from datetime import *;
from contextlib import suppress;
import time;
from time import strftime;
import os;
import datetime as dt;
from ctypes import create_unicode_buffer, windll;
import xlwt;
import win32com.client;


class ResumoExecutivo:
    
    def __init__(self, sistema, df, path, plan_dados, pasta_cod):
        # recebe como parametro o sistema em que estao as informacoes e o problema com o modelo
        self.sin = sistema;
        self.df = df;
        self.caminho = path;
        self.planilha = plan_dados;
        self.vMDI = pasta_cod;

        return;

    def imprime(self, dfUHE, dfCustos):
        
        # chama os metodos para imprimir as saidas
        #with suppress(Exception):
        self.imprimeResumoExecutivo(dfUHE, dfCustos);
        
        return;


    def imprimeResumoExecutivo(self, dfUHE, dfCustos):
        # abre o arquivo e pega a aba
        wb = load_workbook(self.caminho + "Resumo.xlsx");  	
        aba = wb.get_sheet_by_name("Resumo Executivo");
		
        # pega objetos basicos
        df = self.df;
        sin = self.sin;
        
        # pega o ano inicial da impressao
        anoInicial = int(aba.cell(row=7,column=5).value - self.sin.anoInicial);
        ncols = self.sin.numAnos-anoInicial;
        
        # vetor para armazenar totais
        total = [0] * (ncols+1);      
        
        # percorre as linhas para saber o tipo de fonte
        linha = 0;
        hdic = {};
        while (aba.cell(row=8,column=23).offset(row=linha).value is not None):
            # pega o tipo
            tipoProj = aba.cell(row=8,column=23).offset(row=linha).value;
            
            # pega os projetos
            projs = aba.cell(row=8,column=24).offset(row=linha).value;
            
            # prepara a lista de projetos
            lista_projs = [];
            if (projs != "") and (projs is not None):
                projs = str(projs);
                lista_projs = projs.split(";");
                # converte para inteiro
                lista_projs = [int(float(item)) for item in lista_projs];
            
            # inicializa o valor acumulado
            val_acum = [0];
            val_inv = 0;
            ano_PDE = 10;
            val_inv_PDE = 0;
			

            with suppress(Exception):
				# percorre os anos
                for iano in range(anoInicial,self.sin.numAnos):
					# pega o iper como dezembro do ano
                    iper = int((iano+1)*12-1);
					
					# verifica o tipo de projetos
                    if tipoProj == "Hidro":
						# percorre todos os projetos
                        val = 0;
                        for hidro in list(sin.listaGeralProjUHE.values()):
							# pega o registro referente a esta UHE na saida de expansao
                            row = dfUHE.loc[dfUHE["codNW"]==hidro.indexUsinaExterno]

							# caso tenha retornado algum valor
                            if len(row) > 0:
								# pega o primeiro e unico elemento
                                row = row.iloc[0]
                                hdic[hidro.nomeUsina]=row["iper"]-1;

								# verifica se o periodo que expandiu eh inferior
                                if (row["iper"]-1) <= iper:
									# soma o investimento
                                    val_inv += hidro.custoFixo*min(12,iper-(row["iper"]-1))
                                    val += hidro.potUsina;
                            else:
                                hdic[hidro.nomeUsina]=999;

                        val_acum.append(val);
					
                    if tipoProj == "Reversivel":
						# pega o valor do periodo de todos os projetos
                        val_acum.append(sum(df.loc[iper][sin.listaIndGeralProjReversivel[revers-1].nomeUsina] for revers in lista_projs));
                        val_inv += sum(sin.listaIndGeralProjReversivel[revers-1].custoMensal*df.loc[iper_all][sin.listaIndGeralProjReversivel[revers-1].nomeUsina] for revers in lista_projs for iper_all in range(iano*12, (iano+1)*12));
					
                    if tipoProj == "RenovCont":
						# pega o valor do periodo de todos os projetos
                        val_acum.append(sum(df.loc[iper][sin.listaIndGeralProjRenov[renov-1].nomeUsina] for renov in lista_projs));
						# sin.listaIndGeralProjRenov[renov-1].nomeUsina porque tem pegar somente as usinas cujo indice estao dentro de lista_projs
                        val_inv += sum(sin.listaIndGeralProjRenov[renov-1].custoMensal*df.loc[iper_all][sin.listaIndGeralProjRenov[renov-1].nomeUsina] for renov in lista_projs for iper_all in range(iano*12, (iano+1)*12));

                    if tipoProj == "TermCont":
						# pega o valor do periodo de todos os projetos
                        val_acum.append(sum(df.loc[iper][sin.listaIndGeralProjTerm[term].nomeUsina.replace(";",",")]/sin.listaIndGeralProjTerm[term].fdisp for term in lista_projs));
						# sin.listaIndGeralProjRenov[term].nomeUsina porque tem pegar somente as usinas cujo indice estao dentro de lista_projs
                        val_inv += sum(sin.listaIndGeralProjTerm[term].custoFixo*df.loc[iper_all][sin.listaIndGeralProjTerm[term].nomeUsina.replace(";",",")] for term in lista_projs for iper_all in range(iano*12, (iano+1)*12));

                    if tipoProj == "TermContT":
						# pega o valor do periodo de todos os projetos
                        val_acum.append(sum(df.loc[iper][sin.listaIndGeralProjTerm[term].nomeUsina.replace(";",",")]/sin.listaIndGeralProjTerm[term].fdisp for term in lista_projs));
						# sin.listaIndGeralProjRenov[term].nomeUsina porque tem pegar somente as usinas cujo indice estao dentro de lista_projs
                        val_inv += sum(sin.listaIndGeralProjTerm[term].custoFixo*df.loc[iper_all][sin.listaIndGeralProjTerm[term].nomeUsina.replace(";",",")] for term in lista_projs for iper_all in range(iano*12, (iano+1)*12));

                    if tipoProj == "TermInt":
						# pega o valor do periodo de todos os projetos
                        val_acum.append(sum(df.loc[iper][sin.listaIndGeralProjTerm[term].nomeUsina]/sin.listaIndGeralProjTerm[term].fdisp for term in lista_projs));
                        val_inv += sum(sin.listaIndGeralProjTerm[term].custoFixo*df.loc[iper_all][sin.listaIndGeralProjTerm[term].nomeUsina] for term in lista_projs for iper_all in range(iano*12, (iano+1)*12));

                    # escreve o valor incremental
                    col = iano-anoInicial;
                    val_inc = val_acum[col+1]-val_acum[col];
                    aba.cell(row=8,column=5).offset(row=linha,column=col).value = val_inc;
                    
                    if tipoProj != "TermContT":
                    # soma no total geral
                        total[col] += val_inc;
                        total[ncols] += val_inc;
            
                # imprime o total da fonte
                aba.cell(row=8,column=5).offset(row=linha,column=col+1).value = val_acum[col+1];

                # imprime o total para 2029
                aba.cell(row=8,column=5).offset(row=linha,column=col+2).value = val_acum[9];

                # imprime o total de investimento - divide por 1 milhao por definicao de unidade na tabela do excel
                aba.cell(row=8,column=5).offset(row=linha,column=col+4).value = val_inv/1000000;

                # imprime o total de investimento no ano alvo - divide por 1 milhao por definicao de unidade na tabela do excel
                aba.cell(row=8,column=5).offset(row=linha,column=col+5).value = val_inv_PDE/1000000;
            
            # proxima linha
            linha += 1;

        # insere um metadado referente ao numero de linhas
        aba.cell(row=1,column=1).value = linha+1;
                    
        # imprime os totais gerais anuais
        for col in range(ncols+1): 
            aba.cell(row=8,column=5).offset(row=linha,column=col).value = total[col];
            
        # guarda as informacoes
        (self.nlinhas, self.ncols) = (linha,ncols);

        # imprime o caminho
        aba.cell(row=1,column=2).value = self.caminho
        
        # imprime a saida de UHE novas
        self.imprimeResumoUHE(aba, self.nlinhas+8, hdic);

        # fecha a planilha de resumo executivo
        wb.save(self.caminho + "Resumo.xlsx");

        return;
        
    def imprimeResumoUHE(self, aba, linha_inicio, hdic):
        nmeses = self.sin.numMeses;

        # limpa celulas
        for row in aba['D'+str(self.nlinhas+15)+':G200']:
            for cell in row:
                cell.value = None

        nomeSubsis = ["SUDESTE","SUL","NORDESTE","NORTE","ITAIPU","AC RO","MAN/AP/BV","B.MONTE","T. PIRES","PARANA","TAPAJOS","IVAIPORA","IMPERATRIZ","XINGU"];

        # monta um dicionario com os nomes e os periodos
        
        listaUHE = {hidro.nomeUsina: hdic[hidro.nomeUsina] for hidro in list(self.sin.listaGeralProjUHE.values())};
        
        # ordena a lista pelo periodo de entrada
        pos = linha_inicio;
        for nome in sorted(listaUHE, key=listaUHE.get):
            # imprime apenas aquelas que entrar no horizonte
            per = listaUHE[nome];
            if per < nmeses:
                # pega o projeto
                proj = self.sin.listaGeralProjUHE[nome];
                aba.cell(row=8,column=4).offset(row=pos).value = nome;
                aba.cell(row=8,column=5).offset(row=pos).value = nomeSubsis[int(proj.sis_index)-1];
                aba.cell(row=8,column=6).offset(row=pos).value = proj.potUsina;
                aba.cell(row=8,column=7).offset(row=pos).value = str(int(per%12+1)) + "/" + str(int(int(self.sin.anoInicial)+per//12));
                pos += 1
                
        return;

    def imprimeResumoTransmissao(self, aba, linha_inicio):
        m = self.modelo;
        nmeses = self.sin.numMeses;

        # limpa celulas
        for row in aba['I50:L72']:
            for cell in row:
                cell.value = None

        nomeSubsis = ["SUDESTE","SUL","NORDESTE","NORTE","ITAIPU","AC RO","MAN/AP/BV","B.MONTE","T. PIRES","PARANA","TAPAJOS","IVAIPORA","IMPERATRIZ","XINGU"];

        pos = linha_inicio;
        for iper in range(1,nmeses-1):
            for isis in m.subsistemas:
                for jsis in m.subsistemas:
                    if m.capExpInter[isis,jsis,iper].value > m.capExpInter[isis,jsis,iper-1].value:
                        aba.cell(row=8,column=9).offset(row=pos).value = nomeSubsis[isis];
                        aba.cell(row=8,column=10).offset(row=pos).value = nomeSubsis[jsis];
                        aba.cell(row=8,column=11).offset(row=pos).value = m.capExpInter[isis,jsis,iper].value - m.capExpInter[isis,jsis,iper-1].value;
                        aba.cell(row=8,column=12).offset(row=pos).value = str(int(iper%12+1)) + "/" + str(int(int(self.sin.anoInicial)+iper//12));
                        pos += 1
                
        return;
		
    def imprimeSistema(self, df):
        sin = self.sin;
        
        # abre o arquivo para a saidas
        saidaResul = open(self.caminho + "sistemaRev.dat", "w");

        # imprime o cabecalho
        saidaResul.write(" MERCADO DE ENERGIA TOTAL\n");
        saidaResul.write(" XXX\n");
        saidaResul.write("       XXXJAN. XXXFEV. XXXMAR. XXXABR. XXXMAI. XXXJUN. XXXJUL. XXXAGO. XXXSET. XXXOUT. XXXNOV. XXXDEZ.\n");

        for isis in range(0,sin.nsis-3): #nao imprime os ficticios
            subs = self.sin.subsistemas[isis]
            saidaResul.write('{0:>4s}'.format(str(isis+1))+"\n");
            for iano in range(0,sin.numAnos+1):
                if iano != sin.numAnos:
                    linha = str(sin.anoInicial+iano) + "  ";
                    for imes in range(12):
                        #verifica se tem reversivel neste subsistema
                        cap_rev = 0;
                        cap_rev = sum([sum(df.loc[df["iper"]==iano*12+imes, rev.nomeUsina].values)/(1/rev.rendimento-1) for rev in subs.listaProjReversivel])
                        demanda_adic = cap_rev*sum(sin.duracaoPatamar[ipat][iano*12+imes] for ipat in sin.naoBombReversivel);
                        linha += '%7d' % (round(subs.demandaMedia[iano*12+imes] + demanda_adic)) + ".";
                else:
                    linha = linha.replace(str(sin.anoInicial + sin.numAnos -1), "POS ");
                saidaResul.write(linha + "\n");
        return;	
